from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Create a new Workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

# Write "Hello world" in uppercase to cell A1
# ws['C4'] = "Hello world"
# Write "Hello world" in lowercase to cell A2

# Save the workbook
wb.save("monitoring.xlsx")
