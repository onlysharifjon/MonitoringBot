import distutils.msvccompiler

from aiogram.dispatcher import FSMContext

from bot import dp, ADMINS
from aiogram import types

from database import connect, cursor


@dp.message_handler(text="Xodimlar ro`yxati")
async def xodim_royxat(message: types.Message):
    malumotlar = cursor.execute("SELECT * FROM xodimlar").fetchall()
    txt = ""
    for i in malumotlar:
        txt += f"{i[0]})👨‍💻Xodim: <b>{i[3]}</b>\n\n🆔-> <code>{i[1]}</code>\n\n"
    await message.answer(txt)


from bot import Shogirdchalar


@dp.message_handler(text='Xodim qo`shish')
async def xodim_qoshish(message: types):
    if str(message.from_user.id) in ADMINS:
        await message.answer('Xodim ismini kiriting ⏬ !')
        await Shogirdchalar.new_xodim_name.set()
    else:
        await message.answer('Siz admin emassiz ❌')


@dp.message_handler(state=Shogirdchalar.new_xodim_name)
async def new_xodim_name(message: types.Message):
    global ismi
    ismi = message.text
    await message.answer('Xodim ID sini kiriting ⏬ !')
    await Shogirdchalar.new_xodim_id.set()


@dp.message_handler(state=Shogirdchalar.new_xodim_id)
async def new_xodim_id(message: types.Message, state: FSMContext):
    id_user = message.text
    vaqt = message.date
    cursor.execute('INSERT INTO xodimlar(user_id, time, name) VALUES (?,?,?)', (id_user, vaqt, ismi))
    await message.answer(f'{ismi} Bazaga qo`shildi ✅')
    await state.finish()


@dp.message_handler(text='Xodimni o`chirish')
async def xodim_delete(message: types.Message):
    if str(message.from_user.id) in ADMINS:
        await message.answer('Xodim id ma`lumotini kiriting ⏬ !')
        await Shogirdchalar.xodim_ochirish.set()
    else:
        await message.answer('Siz admin emassiz ❌')


@dp.message_handler(state=Shogirdchalar.xodim_ochirish)
async def deleter(message: types.Message, state: FSMContext):
    xodim_id = message.text
    try:
        xodim_id = int(xodim_id)
        cursor.execute(f'DELETE FROM xodimlar WHERE user_id={xodim_id}')
        connect.commit()
        await message.answer('Xodim bazadan o`chirildi ✅')
    except:
        await message.answer('Siz xato id kiritdingiz!😒')


@dp.message_handler(text="📝 1 kunlik ma`lumot")
async def day_1(message: types.Message):
    day = str(message.date.day)
    month = str(message.date.month)
    year = str(message.date.year)

    check = cursor.execute("SELECT * FROM monitoring WHERE keldi_kun=? AND oy=? AND yil=?", (
        day, month, year)).fetchall()
    txt = ""
    txt += f'⏳{year}-{month}-{day}\n\n'
    for i in check:
        txt += f"<b>{i[1]}</b>🆔: Keldi: <i>{i[4]}</i> Ketdi: <i>{i[6]}</i>\n\n"

    await message.answer(txt)


from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup


@dp.message_handler(text='📝 7 kunlik ma`lumot')
async def kun7_monitoring(message: types.Message):
    if str(message.from_user.id) in ADMINS:
        xodimchalar = cursor.execute('SELECT name FROM xodimlar').fetchall()

        xodimlar_inline_button = InlineKeyboardMarkup()

        for i in xodimchalar:
            xodimlar_inline_button.add(InlineKeyboardButton(text=f"{i[0]}", callback_data=f"{i[0]}"))
        await message.answer('⏬   Xodimlar   ⏬', reply_markup=xodimlar_inline_button)
    else:
        await message.answer("Siz admin emassiz ❌")


@dp.callback_query_handler()
async def xodim_nomi(call: types.CallbackQuery):
    await call.message.delete()
    bugunki_kun = int(call.message.date.day)

    xodim_id = cursor.execute(f'SELECT user_id FROM xodimlar WHERE name=?', (call.data,)).fetchone()
    xodim_id = xodim_id[0]
    baza = []
    for i in range(7):
        filtr_day = bugunki_kun - i
        kun_7 = cursor.execute("SELECT * FROM monitoring WHERE user_id=? AND keldi_kun=?",
                               (xodim_id, filtr_day)).fetchall()
        baza.append(kun_7)
    txt = ""
    count = 0

    for i in baza:

        print(i)
        if i:
            count += 1
            txt += f"⏳<b>{i[0][9]}-{i[0][8]}-{i[0][5]} {call.data}</b>\n🆔: {i[0][1]}\n📍 {i[0][2], i[0][3]}\n➡️ Keldi ⌚️: {i[0][4]}\n⬅️ Ketdi ⌚️: {i[0][6]}\n\n"
        else:
            txt += f'Sana: {bugunki_kun - count} kelmagan❌\n\n'
            count += 1
    await call.message.answer(txt)
    # txt = ""
    # for i in kun_7:
    #     txt+=f"\n\n⏳{i[9]}-{i[8]}-{i[7]}____{call.data}_____\n\n🆔: {i[1]}\n📍 {i[2],i[3]}\n➡️ Keldi ⌚️: {i[4]}\n⬅️ Ketdi ⌚️: {i[6]}"
    # await call.message.answer(txt)


#

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

# Create a new Workbook
wb = Workbook()


@dp.message_handler(text='📝 Oy ma`lumoti')
async def month_data(message: types.Message):
    oy = cursor.execute('SELECT * FROM monitoring').fetchall()
    idlar = []
    for i in oy:
        if i[1] not in idlar:
            idlar.append(i[1])

    xodimlar_ismlari = []
    for i in idlar:
        xodim_ismi = cursor.execute('SELECT name FROM xodimlar WHERE user_id=?', (i,)).fetchone()
        if xodim_ismi:
            xodimlar_ismlari.append(xodim_ismi[0])

    # Select the active worksheet
    ws = wb.active
    cell = ws

    # Write "Hello world" in uppercase to cell A1
    count: int = 0
    for i in xodimlar_ismlari:
        xodim_id = cursor.execute('SELECT user_id FROM xodimlar WHERE name=?', (i,)).fetchone()
        xodim_id = xodim_id[0]
        time_user = cursor.execute('SELECT * FROM monitoring WHERE user_id=?', (xodim_id,)).fetchall()

        print(f"Xodim nomi: {i} vaqtari: {time_user}")
        count += 1
        for b in time_user:
            ws.cell(row=count, column=b[5]+1).value = f"Keldi_kun: {b[5]}--{b[4]}// Ketdi_kun: {b[7]}--{b[6]}"
            # Add background color green if there is some information
            if b[4] or b[6]:
                thin_border = Border(left=Side(style='medium', color='000000'),
                                     right=Side(style='medium', color='000000'),
                                     top=Side(style='medium', color='000000'),
                                     bottom=Side(style='medium', color='000000'))
                ws.cell(row=count, column=b[5] + 1).border = thin_border
                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                ws.cell(row=count, column=b[5]+1).fill = green_fill

        ws.cell(row=count, column=1).value = i
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws.cell(row=count, column=1).fill = red_fill

        thin_border = Border(left=Side(style='medium', color='000000'),
                             right=Side(style='medium', color='000000'),
                             top=Side(style='medium', color='000000'),
                             bottom=Side(style='medium', color='000000'))
        ws.cell(row=count, column=1).border = thin_border

        list_alphabet = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                         'U', 'V', 'W', 'X', 'Y', 'Z']
        for k in list_alphabet:
            ws.column_dimensions[k].width = 42
        ws.column_dimensions['A'].width = 30


    # Fill empty cells with pink color
    max_row = ws.max_row
    max_column = ws.max_column
    for row in range(1, max_row+1):
        for column in range(1, max_column+1):

            if not ws.cell(row=row, column=column).value:
                pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
                ws.cell(row=row, column=column).fill = pink_fill
                thin_border = Border(left=Side(style='medium', color='000000'),
                                     right=Side(style='medium', color='000000'),
                                     top=Side(style='medium', color='000000'),
                                     bottom=Side(style='medium', color='000000'))
                ws.cell(row=row, column=column).border = thin_border

    wb.save("monitoring.xlsx")

    #excel file ni jo`natish
    await message.answer_document(open('monitoring.xlsx','rb'),caption='Xodimlarning 1 oylik kelgan va ketgan ma`lumotlari')
from keyboards.default import oylik
@dp.message_handler(text='💸Oylik hisoboti')
async def oylik_funcsion(message:types.Message):
    print(ADMINS)
    print(message.from_user.id)
    if str(message.from_user.id) in ADMINS:

        print(True)
        await message.answer("Menyu yangilandi:)",reply_markup=oylik)


from bot import *

@dp.message_handler(text='🔙orqaga')
async def back(message: types.Message):
    await message.answer('Orqaga qaytildi',reply_markup=admin)

